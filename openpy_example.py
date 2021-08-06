from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename = './data/b1.xlsx')
sheet_ranges = wb['Total']
print(sheet_ranges['J7'].value)
sheet_ranges['J5'] = 56

for x in range(1,100):
    print('E' + str(x))
    print(sheet_ranges['E' + str(x)].value)

#wb2 = Workbook()

# grab the active worksheet
#ws2 = wb2.active

# Data can be assigned directly to cells
#ws2['J7'] = 42

# Rows can also be appended
#ws2.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws2['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")